"""Load the operational workbook into SQLite.

Two decisions matter here.

**The snapshot time is the clock.** The workbook states when it was taken, and
every SLA breach, ticket age and pickup delay must be measured against that
instant rather than against today. Reading the real clock would make answers drift
the moment the data stops being fresh -- and the pack is already dated.

**Timestamps are stored with an offset.** The sheets hold naive local times like
``2026-08-16 09:00``. They are read in the snapshot's timezone and written back as
ISO-8601 with an offset, so nothing downstream has to guess what they mean.

Sheets are mapped to tables generically: column names and types come from the
data, so a workbook with different sheets loads without code changes.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from datetime import date as date_type
from datetime import datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

METADATA_SHEET_HINT = "readme"
METADATA_TABLE = "corpus_meta"
SNAPSHOT_KEY = "snapshot_at"

# "2026-08-16 11:00 Asia/Kolkata" -- an optional IANA zone after a naive timestamp.
_SNAPSHOT = re.compile(
    r"(?P<date>\d{4}-\d{2}-\d{2})[ T](?P<time>\d{2}:\d{2}(?::\d{2})?)"
    r"(?:\s+(?P<zone>[A-Za-z]+(?:/[A-Za-z_+-]+)+|UTC))?"
)
_NAIVE_TIMESTAMP = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(?::\d{2})?$")
_IDENTIFIER_NOISE = re.compile(r"[^0-9a-z_]+")

# A workbook that states no timezone is read as UTC, so results stay reproducible
# rather than depending on the machine that ran ingest.
FALLBACK_TIMEZONE = ZoneInfo("UTC")


@dataclass(frozen=True)
class Workbook:
    """Sheets as rows of plain values, plus the metadata that dates them."""

    sheets: dict[str, list[dict[str, Any]]]
    metadata: dict[str, str]
    snapshot_at: datetime

    def table(self, name: str) -> list[dict[str, Any]]:
        return self.sheets[name]


def read_workbook(path: Path) -> Workbook:
    """Read every sheet, resolving timestamps against the stated snapshot time."""
    raw = load_workbook(path, data_only=True, read_only=True)
    try:
        metadata: dict[str, str] = {}
        tables: dict[str, list[list[Any]]] = {}
        headers: dict[str, list[str]] = {}

        for name in raw.sheetnames:
            rows = [list(row) for row in raw[name].iter_rows(values_only=True)]
            rows = [row for row in rows if any(cell is not None for cell in row)]
            if not rows:
                continue
            if METADATA_SHEET_HINT in name.strip().lower():
                metadata.update(_as_key_values(rows))
                continue
            headers[name] = [_identifier(cell) for cell in rows[0]]
            tables[name] = rows[1:]

        snapshot_at = _snapshot_time(metadata)
        sheets = {
            _identifier(name): _rows_to_dicts(headers[name], body, snapshot_at.tzinfo)
            for name, body in tables.items()
        }
        return Workbook(sheets=sheets, metadata=metadata, snapshot_at=snapshot_at)
    finally:
        raw.close()


def write_database(workbook: Workbook, db_path: Path) -> None:
    """Write the workbook to a fresh SQLite database at ``db_path``."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.unlink(missing_ok=True)

    connection = sqlite3.connect(db_path)
    try:
        for table, rows in workbook.sheets.items():
            _create_table(connection, table, rows)
        _write_metadata(connection, workbook)
        connection.commit()
    finally:
        connection.close()


def build_database(source: Path, db_path: Path) -> Workbook:
    workbook = read_workbook(source)
    write_database(workbook, db_path)
    return workbook


def _snapshot_time(metadata: dict[str, str]) -> datetime:
    """Find the snapshot instant among the metadata rows.

    The key is matched loosely -- any metadata row mentioning a snapshot will do --
    so the workbook can reword its own README without breaking ingest.
    """
    for key, value in metadata.items():
        if "snapshot" not in key.lower():
            continue
        match = _SNAPSHOT.search(str(value))
        if match:
            return _to_aware(match)
    raise ValueError(
        "the workbook states no dataset snapshot time; every time-based answer "
        "depends on it, so ingest cannot continue"
    )


def _to_aware(match: re.Match[str]) -> datetime:
    stamp = f"{match.group('date')} {match.group('time')}"
    fmt = "%Y-%m-%d %H:%M:%S" if stamp.count(":") == 2 else "%Y-%m-%d %H:%M"
    zone = match.group("zone")
    tzinfo = ZoneInfo(zone) if zone else FALLBACK_TIMEZONE
    return datetime.strptime(stamp, fmt).replace(tzinfo=tzinfo)


def _as_key_values(rows: list[list[Any]]) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for row in rows:
        if len(row) >= 2 and row[0] is not None and row[1] is not None:
            pairs[str(row[0]).strip()] = str(row[1]).strip()
    return pairs


def _rows_to_dicts(
    header: list[str], body: list[list[Any]], tzinfo: Any
) -> list[dict[str, Any]]:
    return [
        {
            column: _normalise(value, tzinfo)
            for column, value in zip(header, row, strict=False)
            if column
        }
        for row in body
    ]


def _normalise(value: Any, tzinfo: Any) -> Any:
    """Give every value a form SQLite can store and a reader can interpret."""
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, datetime):
        return _attach(value, tzinfo).isoformat()
    if isinstance(value, date_type):
        return _attach(datetime.combine(value, time.min), tzinfo).isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if _NAIVE_TIMESTAMP.match(text):
            return _attach(datetime.fromisoformat(text), tzinfo).isoformat()
        return text
    return value


def _attach(moment: datetime, tzinfo: Any) -> datetime:
    """Read a naive timestamp as local to the snapshot's timezone."""
    return moment if moment.tzinfo else moment.replace(tzinfo=tzinfo)


def _create_table(connection: sqlite3.Connection, table: str, rows: list[dict[str, Any]]) -> None:
    columns = _columns(rows)
    if not columns:
        return
    definitions = ", ".join(f'"{name}" {sql_type}' for name, sql_type in columns.items())
    connection.execute(f'CREATE TABLE "{table}" ({definitions})')
    placeholders = ", ".join("?" for _ in columns)
    names = ", ".join(f'"{name}"' for name in columns)
    connection.executemany(
        f'INSERT INTO "{table}" ({names}) VALUES ({placeholders})',
        [tuple(row.get(name) for name in columns) for row in rows],
    )


def _columns(rows: list[dict[str, Any]]) -> dict[str, str]:
    """Infer a column order and type from the rows themselves."""
    ordered: dict[str, None] = {}
    for row in rows:
        ordered.update(dict.fromkeys(row))
    return {name: _sql_type([row.get(name) for row in rows]) for name in ordered}


def _sql_type(values: list[Any]) -> str:
    present = [value for value in values if value is not None]
    if present and all(isinstance(value, int) for value in present):
        return "INTEGER"
    if present and all(isinstance(value, int | float) for value in present):
        return "REAL"
    return "TEXT"


def _write_metadata(connection: sqlite3.Connection, workbook: Workbook) -> None:
    connection.execute(f'CREATE TABLE "{METADATA_TABLE}" (key TEXT PRIMARY KEY, value TEXT)')
    entries = {SNAPSHOT_KEY: workbook.snapshot_at.isoformat()}
    entries.update({_identifier(key): value for key, value in workbook.metadata.items()})
    connection.executemany(
        f'INSERT OR REPLACE INTO "{METADATA_TABLE}" (key, value) VALUES (?, ?)',
        list(entries.items()),
    )


def _identifier(value: Any) -> str:
    """Turn a sheet or column label into a safe snake_case SQL identifier."""
    slug = _IDENTIFIER_NOISE.sub("_", str(value or "").strip().lower()).strip("_")
    return f"_{slug}" if slug[:1].isdigit() else slug
